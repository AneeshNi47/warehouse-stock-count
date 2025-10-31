from flask import Blueprint, flash, redirect, jsonify, render_template, request, url_for
from flask_login import login_required, current_user
from app.models import Location, Warehouse, User, ScanLine, ScanRecord
from app import db
from app.constants.status import ScanLineStatus
from flask import send_file
import io
from app.utils.s3_helper import generate_presigned_url
from openpyxl import Workbook
from datetime import datetime



bp = Blueprint("team_leader", __name__, url_prefix="/teamleader")


@bp.route("/dashboard")
@login_required
def dashboard():
    locations = Location.query.all()
    statuses = [
        ScanLineStatus.CREATED,
        ScanLineStatus.ALLOCATED,
        ScanLineStatus.IN_PROGRESS,
        ScanLineStatus.VARIANCE_APPROVED,
        ScanLineStatus.COMPLETED,
        ScanLineStatus.DISCARDED,
        ScanLineStatus.VARIATION_COUNT_COMPLETED,
        ScanLineStatus.VARIATION_ADDITIONAL_REQUIRED
    ]
    warehouses = [w.to_dict() for w in Warehouse.query.all()]
    counters = User.query.filter_by(role="Counter", is_active=True).all()

    my_lines = ScanLine.query.filter_by(team_leader=current_user).all()
    other_lines = ScanLine.query.filter(ScanLine.team_leader != current_user).all()

    return render_template(
        "team_leader_dashboard.html",
        locations=locations,
        warehouses=warehouses,
        counters=counters,
        my_lines=my_lines,
        other_lines=other_lines,
        statuses=statuses
    )


@bp.route("/create-scan-line", methods=["POST"])
@login_required
def create_scan_line():
    # Only Team Leader can create scan lines
    if current_user.role != "TeamLeader":
        flash("Access denied", "danger")
        return redirect(url_for("team_leader.dashboard"))

    location_id = request.form.get("location_id")
    warehouse_id = request.form.get("warehouse_id")
    target_count = request.form.get("target_count")
    counter_1_id = request.form.get("counter_1_id")
    counter_2_id = request.form.get("counter_2_id")

    # Generate a line code (you can improve this logic later)
    from datetime import datetime
    line_code = f"LINE-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"

    scan_line = ScanLine(
        line_code=line_code,
        location_id=location_id,
        warehouse_id=warehouse_id,
        target_count=target_count,
        counter_1_id=counter_1_id,
        counter_2_id=counter_2_id,
        team_leader=current_user
    )

    db.session.add(scan_line)
    db.session.commit()

    flash("New scan line created successfully", "success")
    return redirect(url_for("team_leader.dashboard"))


@bp.route('/scan_line/<int:id>')
@login_required
def view_scan_line(id):
    line = ScanLine.query.get_or_404(id)

    records = ScanRecord.query.filter_by(scan_line_id=line.id).all()
    # Generate presigned URLs for each record image
    for record in records:
        if record.image_path:
            record.image_url = generate_presigned_url(record.image_path)
        else:
            record.image_url = None
    return render_template('team_leader_view_scan_line.html', line=line, scan_records=records)


@bp.route('/scan_line/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_scan_line(id):
    line = ScanLine.query.get_or_404(id)

    if line.team_leader_user_id != current_user.id:
        flash("You don't have permission to edit this scan line.", "danger")
        return redirect(url_for('team_leader.dashboard'))

    locations = Location.query.all()
    warehouses = Warehouse.query.all()
    counters = User.query.filter_by(role='Counter', is_active=True).all()

    if request.method == 'POST':
        line.location_id = request.form.get('location_id')
        line.warehouse_id = request.form.get('warehouse_id')
        line.target_count = request.form.get('target_count')
        line.counter_1_id = request.form.get('counter_1_id')
        line.counter_2_id = request.form.get('counter_2_id')

        db.session.commit()
        flash("Scan line updated successfully!", "success")
        return redirect(url_for('team_leader.dashboard'))

    return render_template(
        'team_leader_edit_scan_line.html',
        line=line,
        locations=locations,
        warehouses=warehouses,
        counters=counters
    )


@bp.route('/scan_line/<int:id>/delete')
@login_required
def delete_scan_line(id):
    line = ScanLine.query.get_or_404(id)

    if line.team_leader_user_id != current_user.id:
        flash("You don't have permission to delete this scan line.", "danger")
        return redirect(url_for('team_leader.dashboard'))

    db.session.delete(line)
    db.session.commit()
    flash("Scan line deleted successfully!", "success")
    return redirect(url_for('team_leader.dashboard'))


@bp.route('/approve_variation', methods=['POST'])
@login_required
def approve_variation():
    """Team Leader approves variation requests (count completed / additional required)."""
    line_id = request.form.get("line_id")
    action_type = request.form.get("type")
    new_target = request.form.get("new_target_count")

    line = ScanLine.query.get(line_id)
    if not line:
        return jsonify({"success": False, "error": "Scan line not found."}), 404

    # Ensure only the assigned Team Leader can act
    if line.team_leader_user_id != current_user.id:
        return jsonify({"success": False, "error": "Unauthorized action."}), 403

    if action_type == "complete":
        line.status = ScanLineStatus.COMPLETED if hasattr(ScanLineStatus, "COMPLETED") else "Completed"
        line.is_locked = False
        line.target_count = line.current_count
        message = "✅ Line marked as Completed successfully."

    elif action_type == "additional":
        try:
            new_target = int(new_target)
        except (TypeError, ValueError):
            return jsonify({"success": False, "error": "Invalid target count."}), 400

        if new_target <= line.target_count:
            return jsonify({"success": False, "error": "New target must be greater than current count."}), 400

        line.target_count = new_target
        line.is_locked = False
        line.status = ScanLineStatus.IN_PROGRESS if hasattr(ScanLineStatus, "IN_PROGRESS") else "In-Progress"
        message = f"✅ Target count updated to {new_target} and line reopened for counting."

    else:
        return jsonify({"success": False, "error": "Invalid action type."}), 400

    db.session.commit()
    return jsonify({"success": True, "message": message})

@bp.route('/update_scan_record', methods=['POST'])
@login_required
def update_scan_record():
    record_id = request.form.get('record_id')
    record = ScanRecord.query.get(record_id)
    if not record:
        return jsonify({"success": False, "error": "Record not found"})

    record.barcode_1 = request.form.get('barcode_1') or None
    record.barcode_2 = request.form.get('barcode_2') or None
    record.barcode_3 = request.form.get('barcode_3') or None
    db.session.commit()
    return jsonify({"success": True})

@bp.route("/export_custom", methods=["POST"])
@login_required
def export_custom():
    location_ids = request.form.getlist("location_ids")
    warehouse_ids = request.form.getlist("warehouse_ids")
    statuses = request.form.getlist("status_list")

    query = ScanRecord.query.join(ScanLine)

    if location_ids:
        query = query.filter(ScanLine.location_id.in_(location_ids))
    if warehouse_ids:
        query = query.filter(ScanLine.warehouse_id.in_(warehouse_ids))
    if statuses:
        query = query.filter(ScanLine.status.in_(statuses))

    records = query.all()
    if not records:
        flash("No records found for selected filters.", "warning")
        return redirect(url_for("team_leader.dashboard"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Exported Scan Records"

    # Add summary header
    ws["A1"] = "Export Summary"
    ws["A2"] = f"Generated by: {current_user.username}"
    ws["A3"] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A4"] = f"Selected Locations: {', '.join(location_ids) if location_ids else 'All'}"
    ws["A5"] = f"Selected Warehouses: {', '.join(warehouse_ids) if warehouse_ids else 'All'}"
    ws["A6"] = f"Statuses: {', '.join(statuses) if statuses else 'All'}"
    ws["A8"] = "Scan Records:"

    headers = [
        "Location",
        "Warehouse",
        "Line Code",
        "Status",
        "Counter (User)",
        "Partner (Team Leader)",
        "Barcode 1",
        "Barcode 2",
        "Barcode 3",
        "Created Date Time",
    ]
    ws.append([])
    ws.append(headers)

    for r in records:
        line = r.scan_line
        ws.append([
            line.location.name if line.location else "",
            line.warehouse.warehouse_name if line.warehouse else "",
            line.line_code or "",
            line.status or "",
            r.counter_user.username if r.counter_user else "",
            line.team_leader.username if line.team_leader else "",
            r.barcode_1 or "",
            r.barcode_2 or "",
            r.barcode_3 or "",
            r.created_on.strftime("%Y-%m-%d %H:%M:%S"),
        ])

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(c.value)) for c in col if c.value)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"DSV_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )